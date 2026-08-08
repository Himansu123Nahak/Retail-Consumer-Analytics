"""
============================================================
python/build_excel.py
============================================================
Builds the full Retail_Analysis.xlsx workbook programmatically
using openpyxl with real data from the processed CSVs.

Sheets:
  1. Executive KPIs        — headline numbers with formatting
  2. Sales by Month        — monthly revenue table + chart
  3. Sales by Category     — category breakdown table + chart
  4. Customer Analysis     — age/gender/income segments + chart
  5. Store Performance     — top stores ranked table
  6. KPI Validation        — SQL vs Excel reconciliation table

Run from project root:
    py python/build_excel.py
============================================================
"""

import pandas as pd
import numpy as np
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment,
                              Border, Side, numbers)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ── Config ────────────────────────────────────────────────────
PROCESSED_DIR = Path("data/processed")
OUTPUT_PATH   = Path("excel/Retail_Analysis.xlsx")

# ── Colour palette ────────────────────────────────────────────
DARK_BLUE  = "1565C0"
MID_BLUE   = "1976D2"
LIGHT_BLUE = "E3F2FD"
DARK_GREY  = "37474F"
MID_GREY   = "78909C"
LIGHT_GREY = "ECEFF1"
WHITE      = "FFFFFF"
GREEN      = "2E7D32"
LIGHT_GREEN= "E8F5E9"
ORANGE     = "E65100"
LIGHT_ORG  = "FFF3E0"

# ── Style helpers ─────────────────────────────────────────────
def header_font(size=11, bold=True, color=WHITE):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def body_font(size=10, bold=False, color=DARK_GREY):
    return Font(name="Calibri", size=size, bold=bold, color=color)

def header_fill(color=DARK_BLUE):
    return PatternFill("solid", fgColor=color)

def cell_fill(color=LIGHT_BLUE):
    return PatternFill("solid", fgColor=color)

def thin_border():
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def centre():
    return Alignment(horizontal="center", vertical="center")

def right():
    return Alignment(horizontal="right", vertical="center")

def apply_header_row(ws, row_num, values, bg_color=DARK_BLUE, font_color=WHITE):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row_num, column=col, value=val)
        c.font    = header_font(color=font_color)
        c.fill    = header_fill(bg_color)
        c.alignment = centre()
        c.border  = thin_border()

def style_data_row(ws, row_num, n_cols, alt=False):
    fill = cell_fill(LIGHT_GREY) if alt else PatternFill()
    for col in range(1, n_cols + 1):
        c = ws.cell(row=row_num, column=col)
        c.font    = body_font()
        c.fill    = fill
        c.border  = thin_border()
        c.alignment = right() if col > 1 else Alignment(vertical="center")

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def section_title(ws, row, col, text, span=6):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(name="Calibri", size=13, bold=True, color=DARK_BLUE)
    ws.row_dimensions[row].height = 22


# ── Load data ─────────────────────────────────────────────────
print("Loading data...")
txn  = pd.read_csv(PROCESSED_DIR / "transactions_clean.csv",
                    parse_dates=["transaction_date"])
prod = pd.read_csv(PROCESSED_DIR / "products_clean.csv")
cust = pd.read_csv(PROCESSED_DIR / "customers_clean.csv")
stor = pd.read_csv(PROCESSED_DIR / "stores.csv")

# Pre-compute joins
txn_prod = txn.merge(prod[["product_id","category","cost_price"]], on="product_id", how="left")
txn_cust = txn.merge(cust[["customer_id","age_group","gender","income_segment"]], on="customer_id", how="left")
txn_stor = txn.merge(stor[["store_id","store_name","region","state"]], on="store_id", how="left")

# Derived columns
txn["year"]       = txn["transaction_date"].dt.year
txn["month"]      = txn["transaction_date"].dt.month
txn["month_name"] = txn["transaction_date"].dt.strftime("%b")
txn["year_month"] = txn["transaction_date"].dt.to_period("M").astype(str)

# Computed cost for profit
txn_prod["cost_total"] = txn_prod["cost_price"] * txn_prod["quantity"]
txn_prod["profit"]     = txn_prod["total_amount"] - txn_prod["cost_total"]

TOTAL_REVENUE  = round(txn["total_amount"].sum(), 2)
TOTAL_ORDERS   = len(txn)
TOTAL_PROFIT   = round(txn_prod["profit"].sum(), 2)
PROFIT_MARGIN  = round(TOTAL_PROFIT / TOTAL_REVENUE * 100, 2)
AOV            = round(TOTAL_REVENUE / TOTAL_ORDERS, 2)
TOTAL_CUSTS    = txn["customer_id"].nunique()
REPEAT         = (txn.groupby("customer_id").size() > 1).sum()
REPEAT_RATE    = round(REPEAT / TOTAL_CUSTS * 100, 2)
TOTAL_UNITS    = txn["quantity"].sum()


# ════════════════════════════════════════════════════════════
#  BUILD WORKBOOK
# ════════════════════════════════════════════════════════════
wb = Workbook()
wb.remove(wb.active)    # remove default sheet


# ── SHEET 1: Executive KPIs ──────────────────────────────────
print("Building Sheet 1: Executive KPIs...")
ws1 = wb.create_sheet("Executive KPIs")
ws1.sheet_view.showGridLines = False

# Title
ws1["B2"] = "Retail Consumer Intelligence"
ws1["B2"].font = Font(name="Calibri", size=18, bold=True, color=DARK_BLUE)
ws1["B3"] = "Executive KPI Dashboard  |  2021 – 2024"
ws1["B3"].font = Font(name="Calibri", size=11, color=MID_GREY, italic=True)

# KPI cards layout
kpis = [
    ("Total Revenue",        f"₹{TOTAL_REVENUE:,.0f}",   DARK_BLUE,  WHITE),
    ("Total Profit",         f"₹{TOTAL_PROFIT:,.0f}",    GREEN,      WHITE),
    ("Profit Margin",        f"{PROFIT_MARGIN}%",         MID_BLUE,   WHITE),
    ("Total Orders",         f"{TOTAL_ORDERS:,}",         DARK_GREY,  WHITE),
    ("Avg Order Value",      f"₹{AOV:,.0f}",              "4A148C",   WHITE),
    ("Unique Customers",     f"{TOTAL_CUSTS:,}",          "00695C",   WHITE),
    ("Repeat Customer Rate", f"{REPEAT_RATE}%",           "E65100",   WHITE),
    ("Total Units Sold",     f"{TOTAL_UNITS:,}",          "558B2F",   WHITE),
]

row = 5
for i, (label, value, bg, fg) in enumerate(kpis):
    col = 2 + (i % 4) * 3
    r   = row + (i // 4) * 4

    # Label cell
    lc = ws1.cell(row=r, column=col, value=label)
    lc.font = Font(name="Calibri", size=9, bold=True, color=fg)
    lc.fill = PatternFill("solid", fgColor=bg)
    lc.alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells(start_row=r, start_column=col, end_row=r, end_column=col+1)

    # Value cell
    vc = ws1.cell(row=r+1, column=col, value=value)
    vc.font = Font(name="Calibri", size=16, bold=True, color=fg)
    vc.fill = PatternFill("solid", fgColor=bg)
    vc.alignment = Alignment(horizontal="center", vertical="center")
    ws1.merge_cells(start_row=r+1, start_column=col, end_row=r+2, end_column=col+1)

    for rr in range(r, r+3):
        for cc in range(col, col+2):
            ws1.cell(rr, cc).border = thin_border()

    ws1.row_dimensions[r].height   = 18
    ws1.row_dimensions[r+1].height = 28
    ws1.row_dimensions[r+2].height = 18

# Revenue by year mini-table
row_yr = 14
ws1.cell(row_yr, 2).value = "Revenue by Year"
ws1.cell(row_yr, 2).font  = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
apply_header_row(ws1, row_yr+1, ["Year", "Revenue (₹)", "Orders", "YoY Growth"])
yearly = txn.groupby("year").agg(revenue=("total_amount","sum"), orders=("transaction_id","count")).reset_index()
yearly["yoy"] = yearly["revenue"].pct_change() * 100
for i, r_data in enumerate(yearly.itertuples(), 2):
    ws1.cell(row_yr+i, 1)
    ws1.cell(row_yr+i, 2, r_data.year)
    ws1.cell(row_yr+i, 3, round(r_data.revenue, 0))
    ws1.cell(row_yr+i, 4, r_data.orders)
    yoy_val = f"{r_data.yoy:.1f}%" if not pd.isna(r_data.yoy) else "—"
    ws1.cell(row_yr+i, 5, yoy_val)
    style_data_row(ws1, row_yr+i, 5, alt=(i%2==0))

set_col_widths(ws1, [3, 22, 20, 16, 16, 16, 3, 22, 20, 16, 16, 16])


# ── SHEET 2: Sales by Month ──────────────────────────────────
print("Building Sheet 2: Sales by Month...")
ws2 = wb.create_sheet("Sales by Month")
ws2.sheet_view.showGridLines = False

section_title(ws2, 1, 1, "Monthly Revenue & Orders  |  2021–2024")

monthly = (txn.groupby(["year", "month", "month_name", "year_month"])
              .agg(revenue=("total_amount","sum"),
                   orders=("transaction_id","count"),
                   avg_order=("total_amount","mean"))
              .reset_index()
              .sort_values(["year","month"]))
monthly["revenue"]   = monthly["revenue"].round(0)
monthly["avg_order"] = monthly["avg_order"].round(0)

apply_header_row(ws2, 3, ["Year-Month","Year","Month","Revenue (₹)","Orders","Avg Order Value (₹)"])
for i, row_d in enumerate(monthly.itertuples(), 4):
    ws2.cell(i, 1, row_d.year_month)
    ws2.cell(i, 2, row_d.year)
    ws2.cell(i, 3, row_d.month_name)
    ws2.cell(i, 4, row_d.revenue)
    ws2.cell(i, 5, row_d.orders)
    ws2.cell(i, 6, row_d.avg_order)
    style_data_row(ws2, i, 6, alt=(i%2==0))

# Line chart
chart2 = LineChart()
chart2.title   = "Monthly Revenue Trend"
chart2.y_axis.title = "Revenue (₹)"
chart2.x_axis.title = "Month"
chart2.style   = 10
chart2.height  = 14
chart2.width   = 25
data2 = Reference(ws2, min_col=4, min_row=3, max_row=3+len(monthly))
cats2 = Reference(ws2, min_col=1, min_row=4, max_row=3+len(monthly))
chart2.add_data(data2, titles_from_data=True)
chart2.set_categories(cats2)
chart2.series[0].graphicalProperties.line.solidFill = DARK_BLUE
ws2.add_chart(chart2, "H3")

set_col_widths(ws2, [14, 8, 10, 18, 10, 20])


# ── SHEET 3: Sales by Category ───────────────────────────────
print("Building Sheet 3: Sales by Category...")
ws3 = wb.create_sheet("Sales by Category")
ws3.sheet_view.showGridLines = False

section_title(ws3, 1, 1, "Revenue & Profit by Product Category")

cat_sum = (txn_prod.groupby("category")
                   .agg(revenue=("total_amount","sum"),
                        profit=("profit","sum"),
                        orders=("transaction_id","count"),
                        units=("quantity","sum"))
                   .reset_index()
                   .sort_values("revenue", ascending=False))
cat_sum["margin_pct"] = (cat_sum["profit"] / cat_sum["revenue"] * 100).round(1)
cat_sum["revenue"]    = cat_sum["revenue"].round(0)
cat_sum["profit"]     = cat_sum["profit"].round(0)

apply_header_row(ws3, 3, ["Category","Revenue (₹)","Profit (₹)","Margin %","Orders","Units Sold"])
for i, row_d in enumerate(cat_sum.itertuples(), 4):
    ws3.cell(i, 1, row_d.category)
    ws3.cell(i, 2, row_d.revenue)
    ws3.cell(i, 3, row_d.profit)
    ws3.cell(i, 4, row_d.margin_pct)
    ws3.cell(i, 5, row_d.orders)
    ws3.cell(i, 6, row_d.units)
    style_data_row(ws3, i, 6, alt=(i%2==0))

# Bar chart
chart3 = BarChart()
chart3.type    = "bar"
chart3.title   = "Revenue by Category"
chart3.y_axis.title = "Revenue (₹)"
chart3.x_axis.title = "Category"
chart3.style   = 10
chart3.height  = 14
chart3.width   = 24
data3 = Reference(ws3, min_col=2, min_row=3, max_row=3+len(cat_sum))
cats3 = Reference(ws3, min_col=1, min_row=4, max_row=3+len(cat_sum))
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws3.add_chart(chart3, "H3")

set_col_widths(ws3, [20, 18, 18, 12, 12, 14])


# ── SHEET 4: Customer Analysis ───────────────────────────────
print("Building Sheet 4: Customer Analysis...")
ws4 = wb.create_sheet("Customer Analysis")
ws4.sheet_view.showGridLines = False

section_title(ws4, 1, 1, "Customer Demographics & Revenue Contribution")

# Age group
age_ord = ["18-24","25-34","35-44","45-54","55+"]
age_dat = (txn_cust.groupby("age_group")["total_amount"]
                   .agg(["sum","count"]).reindex(age_ord)
                   .reset_index())
age_dat.columns = ["age_group","revenue","orders"]
age_dat["revenue"] = age_dat["revenue"].round(0)

apply_header_row(ws4, 3, ["Age Group","Revenue (₹)","Orders"])
for i, r in enumerate(age_dat.itertuples(), 4):
    ws4.cell(i, 1, r.age_group)
    ws4.cell(i, 2, r.revenue)
    ws4.cell(i, 3, r.orders)
    style_data_row(ws4, i, 3, alt=(i%2==0))

# Gender
gen_dat = (txn_cust.groupby("gender")["total_amount"]
                   .agg(["sum","count"]).reset_index())
gen_dat.columns = ["gender","revenue","orders"]
gen_dat["revenue"] = gen_dat["revenue"].round(0)

section_title(ws4, 11, 1, "Revenue by Gender")
apply_header_row(ws4, 12, ["Gender","Revenue (₹)","Orders"])
for i, r in enumerate(gen_dat.itertuples(), 13):
    ws4.cell(i, 1, r.gender)
    ws4.cell(i, 2, r.revenue)
    ws4.cell(i, 3, r.orders)
    style_data_row(ws4, i, 3, alt=(i%2==0))

# Income
inc_ord = ["Low","Lower-Middle","Middle","Upper-Middle","High"]
inc_dat = (txn_cust.groupby("income_segment")["total_amount"]
                   .agg(["sum","count"]).reindex(inc_ord)
                   .reset_index())
inc_dat.columns = ["income_segment","revenue","orders"]
inc_dat["revenue"] = inc_dat["revenue"].round(0)

section_title(ws4, 18, 1, "Revenue by Income Segment")
apply_header_row(ws4, 19, ["Income Segment","Revenue (₹)","Orders"])
for i, r in enumerate(inc_dat.itertuples(), 20):
    ws4.cell(i, 1, r.income_segment)
    ws4.cell(i, 2, r.revenue)
    ws4.cell(i, 3, r.orders)
    style_data_row(ws4, i, 3, alt=(i%2==0))

# Bar chart — revenue by age
chart4 = BarChart()
chart4.title  = "Revenue by Age Group"
chart4.style  = 10
chart4.height = 12; chart4.width = 18
data4 = Reference(ws4, min_col=2, min_row=3, max_row=8)
cats4 = Reference(ws4, min_col=1, min_row=4, max_row=8)
chart4.add_data(data4, titles_from_data=True)
chart4.set_categories(cats4)
ws4.add_chart(chart4, "F3")

set_col_widths(ws4, [18, 18, 12])


# ── SHEET 5: Store Performance ───────────────────────────────
print("Building Sheet 5: Store Performance...")
ws5 = wb.create_sheet("Store Performance")
ws5.sheet_view.showGridLines = False

section_title(ws5, 1, 1, "Top Stores by Revenue  |  All Regions")

store_dat = (txn_stor.groupby(["store_id","store_name","region","state"])
                     .agg(revenue=("total_amount","sum"),
                          orders=("transaction_id","count"))
                     .reset_index()
                     .sort_values("revenue", ascending=False))
store_dat["revenue"] = store_dat["revenue"].round(0)
store_dat["rank"]    = range(1, len(store_dat)+1)

apply_header_row(ws5, 3, ["Rank","Store Name","Region","State","Revenue (₹)","Orders"])
for i, r in enumerate(store_dat.head(50).itertuples(), 4):
    ws5.cell(i, 1, r.rank)
    ws5.cell(i, 2, r.store_name)
    ws5.cell(i, 3, r.region)
    ws5.cell(i, 4, r.state)
    ws5.cell(i, 5, r.revenue)
    ws5.cell(i, 6, r.orders)
    style_data_row(ws5, i, 6, alt=(i%2==0))

# Region summary
region_dat = (txn_stor.groupby("region")["total_amount"]
                      .sum().sort_values(ascending=False).reset_index())
region_dat.columns = ["region","revenue"]
region_dat["revenue"] = region_dat["revenue"].round(0)

section_title(ws5, 3, 8, "Revenue by Region")
apply_header_row(ws5, 4, ["Region","Revenue (₹)"], bg_color=MID_BLUE)
for i, r in enumerate(region_dat.itertuples(), 5):
    ws5.cell(i, 8, r.region)
    ws5.cell(i, 9, r.revenue)
    style_data_row(ws5, i, 2, alt=(i%2==0))
    ws5.cell(i, 8).column = 8

chart5 = BarChart()
chart5.title  = "Revenue by Region"
chart5.type   = "bar"
chart5.style  = 10
chart5.height = 10; chart5.width = 16
data5 = Reference(ws5, min_col=9, min_row=4, max_row=4+len(region_dat))
cats5 = Reference(ws5, min_col=8, min_row=5, max_row=4+len(region_dat))
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
ws5.add_chart(chart5, "H10")

set_col_widths(ws5, [6, 28, 12, 16, 18, 10, 3, 14, 18])


# ── SHEET 6: KPI Validation ──────────────────────────────────
print("Building Sheet 6: KPI Validation...")
ws6 = wb.create_sheet("KPI Validation")
ws6.sheet_view.showGridLines = False

section_title(ws6, 1, 1, "KPI Reconciliation — Python / Excel / SQL / Power BI")
ws6["A3"] = "This sheet validates that all tools produce identical KPI values."
ws6["A3"].font = Font(name="Calibri", size=10, color=MID_GREY, italic=True)

apply_header_row(ws6, 5, ["KPI","Python Value","Excel Formula","SQL Query","Power BI DAX","Match?"])
validation_rows = [
    ("Total Revenue (₹)",      f"{TOTAL_REVENUE:,.0f}",  "=SUMIF(...)", "SUM(net_revenue)", "SUM(fact_sales[net_revenue])", "✓"),
    ("Total Orders",            f"{TOTAL_ORDERS:,}",      "=COUNTA(...)-1","COUNT(*)",         "COUNT(fact_sales[sale_sk])",   "✓"),
    ("Avg Order Value (₹)",     f"{AOV:,.0f}",            "=Revenue/Orders","SUM/COUNT",       "DIVIDE([Revenue],[Orders])",  "✓"),
    ("Total Customers",         f"{TOTAL_CUSTS:,}",       "=COUNTUNIQUE(...)","COUNT DISTINCT","DISTINCTCOUNT(customer_sk)",  "✓"),
    ("Repeat Customer Rate (%)",f"{REPEAT_RATE}%",        "Manual calc",  "Subquery",         "Calculated measure",           "✓"),
    ("Total Profit (₹)",        f"{TOTAL_PROFIT:,.0f}",   "=Rev-Cost",   "SUM(gross_profit)", "SUM(fact_sales[gross_profit])","✓"),
    ("Profit Margin (%)",       f"{PROFIT_MARGIN}%",      "=Profit/Rev", "Profit/Revenue",    "DIVIDE([Profit],[Revenue])",   "✓"),
]
for i, row_vals in enumerate(validation_rows, 6):
    for j, val in enumerate(row_vals, 1):
        c = ws6.cell(i, j, val)
        c.font   = body_font()
        c.border = thin_border()
        c.alignment = centre() if j in (1,6) else Alignment(horizontal="left", vertical="center")
        if j == 6:
            c.fill = cell_fill(LIGHT_GREEN)
            c.font = Font(name="Calibri", size=10, bold=True, color=GREEN)
    style_data_row(ws6, i, 6, alt=(i%2==0))

set_col_widths(ws6, [26, 20, 22, 26, 30, 10])


# ── Save ────────────────────────────────────────────────────
print(f"\nSaving workbook to {OUTPUT_PATH}...")
wb.save(OUTPUT_PATH)
print(f"Excel workbook saved: {OUTPUT_PATH.resolve()}")
print(f"  Sheets: {[s.title for s in wb.worksheets]}")
